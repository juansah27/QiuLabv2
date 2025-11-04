import pandas as pd
import os
import logging
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .utils import (
    assign_shop_id, convert_date_format, create_gift_id, 
    get_client_id, determine_gift_type, get_marketplace_code
)

class SetupRequestProcessor:
    """Class to handle processing of setup request Excel files"""
    
    def __init__(self, file_path, process_type, created_by, combine_sheets=False, output_format="xlsx"):
        self.file_path = file_path
        self.process_type = process_type
        self.created_by = created_by
        self.combine_sheets = combine_sheets
        self.output_format = output_format
        self.logs = []
        self.output_path = None
        
    def log(self, message, level="info"):
        """Add log message and print to console/log file"""
        log_entry = {"time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
                     "message": message, 
                     "level": level}
        self.logs.append(log_entry)
        
        if level == "error":
            logging.error(message)
        else:
            logging.info(message)
        
        return log_entry
    
    def generate_detailed_error_message(self, error):
        """Generate detailed error message based on error type"""
        error_str = str(error)
        
        # Handle specific error types
        if "MARKETPLACE TIDAK DITEMUKAN DI DATABASE" in error_str:
            return error_str  # Return the original error message
        elif "FORMAT TANGGAL TIDAK VALID" in error_str:
            return error_str  # Return the original error message
        elif "'float' object has no attribute 'lower'" in error_str:
            return self.generate_empty_column_error_message()
        elif "AttributeError" in error_str and "lower" in error_str:
            return self.generate_empty_column_error_message()
        elif "KeyError" in error_str:
            return self.generate_missing_column_error_message(error_str)
        elif "ValueError" in error_str:
            return self.generate_value_error_message(error_str)
        elif error_str.startswith("'") and error_str.endswith("'"):
            # Handle simple KeyError with just column name like "'Client'"
            column_name = error_str.strip("'")
            return f"""❌ KESALAHAN SISTEM

Terjadi kesalahan Kolom '{column_name}' tidak ada

💡 SOLUSI:
• Coba upload file yang memiliki kolom {column_name}
• Hubungi administrator jika masalah berlanjut"""
        else:
            return f"❌ KESALAHAN SISTEM\n\nTerjadi kesalahan yang tidak terduga saat memproses file.\nError: {error_str}\n\n💡 SOLUSI:\n• Coba upload file yang berbeda\n• Hubungi administrator jika masalah berlanjut"
    
    def generate_empty_column_error_message(self):
        """Generate error message for empty column data"""
        return """❌ DATA KOLOM KOSONG

Ditemukan kolom dengan data kosong yang menyebabkan error pemrosesan.

💡 SOLUSI:
• Isi semua kolom wajib yang kosong
• Download template yang benar dari halaman ini"""
    
    def generate_missing_column_error_message(self, error_str):
        """Generate error message for missing columns"""
        return f"""❌ KOLOM TIDAK DITEMUKAN

{error_str}

💡 SOLUSI:
• Download template yang benar dari halaman ini"""
    
    def generate_value_error_message(self, error_str):
        """Generate error message for value errors"""
        return f"""❌ NILAI DATA TIDAK VALID

{error_str}

💡 SOLUSI:
• Periksa format data di file Excel
• Download template yang benar dari halaman ini"""
    
    def validate_required_columns_data(self, df):
        """Validate that required columns have valid data"""
        required_columns = {
            'Marketplace': 'nama marketplace',
            'Client': 'nama client', 
            'ShopId': 'ID toko',
            'SKU Bundle': 'kode bundle',
            'SKU Product': 'kode produk',
            'Qty': 'kuantitas'
        }
        
        empty_columns = []
        
        for col_name, col_desc in required_columns.items():
            if col_name in df.columns:
                # Check for empty values
                empty_mask = df[col_name].isna() | (df[col_name] == '') | (df[col_name].astype(str).str.strip() == '')
                empty_count = empty_mask.sum()
                
                if empty_count > 0:
                    empty_columns.append(f"• Kolom {col_name} ({col_desc}) kosong")
        
        if empty_columns:
            error_msg = f"""❌ DATA KOLOM KOSONG DITEMUKAN

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(empty_columns)}"""
            
            raise ValueError(error_msg)
    
    def validate_supplementary_columns_data(self, df):
        """Validate that required columns for supplementary have valid data"""
        required_columns = {
            'Marketplace': 'nama marketplace',
            'Client': 'nama client',
            'Main SKU': 'kode SKU utama',
            'Gift SKU': 'kode SKU hadiah',
            'Gift Qty': 'kuantitas hadiah',
            'Start_Date': 'tanggal mulai',
            'End_Date': 'tanggal berakhir'
        }
        
        empty_columns = []
        
        for col_name, col_desc in required_columns.items():
            if col_name in df.columns:
                # Check for empty values
                empty_mask = df[col_name].isna() | (df[col_name] == '') | (df[col_name].astype(str).str.strip() == '')
                empty_count = empty_mask.sum()
                
                if empty_count > 0:
                    empty_columns.append(f"• Kolom {col_name} ({col_desc}) kosong")
        
        if empty_columns:
            error_msg = f"""❌ DATA KOLOM SUPPLEMENTARY KOSONG DITEMUKAN

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(empty_columns)}"""
            
            raise ValueError(error_msg)
    
    def validate_gift_columns_data(self, df):
        """Validate that required columns for gift have valid data"""
        required_columns = {
            'Marketplace': 'nama marketplace',
            'Client': 'nama client',
            'Main SKU': 'kode SKU utama',
            'Gift SKU': 'kode SKU hadiah',
            'Gift Qty': 'kuantitas hadiah',
            'Value_Start': 'nilai minimum pembelian',
            'Value_End': 'nilai maksimum pembelian',
            'Start_Date': 'tanggal mulai',
            'End_Date': 'tanggal berakhir'
        }
        
        empty_columns = []
        critical_empty_columns = []
        
        for col_name, col_desc in required_columns.items():
            if col_name in df.columns:
                # Check for empty values
                empty_mask = df[col_name].isna() | (df[col_name] == '') | (df[col_name].astype(str).str.strip() == '')
                empty_count = empty_mask.sum()
                
                if empty_count > 0:
                    # Main SKU is not critical, just log warning
                    if col_name == 'Main SKU':
                        self.log(f"⚠️ Kolom {col_name} ({col_desc}) kosong - Gift type 3 memakai ItemLimit", "warning")
                    else:
                        critical_empty_columns.append(f"• Kolom {col_name} ({col_desc}) kosong")
        
        # Only raise error for critical empty columns (excluding Main SKU)
        if critical_empty_columns:
            error_msg = f"""❌ DATA KOLOM GIFT KOSONG DITEMUKAN

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(critical_empty_columns)}"""
            
            raise ValueError(error_msg)
    
    def validate_shop_id_mapping(self, df):
        """Validate that ShopId mapping is successful"""
        unknown_shop_ids = df[df['ShopId'] == 'Unknown']
        
        if len(unknown_shop_ids) > 0:
            # Get unique marketplace-client combinations that resulted in Unknown
            unknown_combinations = unknown_shop_ids[['Marketplace', 'Client']].drop_duplicates()
            
            error_details = []
            for idx, row in unknown_combinations.iterrows():
                marketplace = row['Marketplace']
                client = row['Client']
                error_details.append(f"• Marketplace: '{marketplace}', Client: '{client}'")
            
            error_msg = f"""❌ DATA KOLOM MARKETPLACE TIDAK DITEMUKAN DI DATABASE

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(error_details)}

💡 SOLUSI:
• Cek Client di Tab Shop Mapping
• Pastikan nama marketplace dan client sesuai dengan yang ada di database
• Periksa ejaan dan format nama marketplace/client"""
            
            raise ValueError(error_msg)
    
    def validate_output_date_columns(self, result_df):
        """Validate that output date columns are not empty"""
        empty_start_dates = result_df[result_df['StartDate'].isna() | (result_df['StartDate'] == '') | (result_df['StartDate'].astype(str).str.strip() == '')]
        empty_end_dates = result_df[result_df['EndDate'].isna() | (result_df['EndDate'] == '') | (result_df['EndDate'].astype(str).str.strip() == '')]
        
        # Check for invalid date format errors
        invalid_start_dates = result_df[result_df['StartDate'].astype(str).str.contains('Format Tanggal Salah', na=False)]
        invalid_end_dates = result_df[result_df['EndDate'].astype(str).str.contains('Format Tanggal Salah', na=False)]
        
        if len(empty_start_dates) > 0 or len(empty_end_dates) > 0 or len(invalid_start_dates) > 0 or len(invalid_end_dates) > 0:
            error_details = []
            
            if len(empty_start_dates) > 0:
                error_details.append(f"• Kolom StartDate: {len(empty_start_dates)} baris kosong")
            
            if len(empty_end_dates) > 0:
                error_details.append(f"• Kolom EndDate: {len(empty_end_dates)} baris kosong")
                
            if len(invalid_start_dates) > 0:
                error_details.append(f"• Kolom StartDate: {len(invalid_start_dates)} baris format tanggal salah")
                
            if len(invalid_end_dates) > 0:
                error_details.append(f"• Kolom EndDate: {len(invalid_end_dates)} baris format tanggal salah")
            
            error_msg = f"""❌ DATA KOLOM TANGGAL BERMASALAH DI HASIL OUTPUT

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(error_details)}

💡 SOLUSI:
• Periksa data tanggal di file Excel asli
• Pastikan format tanggal sesuai template (DD/MM/YYYY atau MM/DD/YYYY)
• Contoh format yang benar: 25/12/2025, 12/25/2025
• Download template yang benar dari halaman ini"""
            
            raise ValueError(error_msg)
    
    def validate_gift_output_date_columns(self, header_df):
        """Validate that gift output date columns are not empty"""
        empty_start_dates = header_df[header_df['STARTDATE'].isna() | (header_df['STARTDATE'] == '') | (header_df['STARTDATE'].astype(str).str.strip() == '')]
        empty_end_dates = header_df[header_df['ENDDATE'].isna() | (header_df['ENDDATE'] == '') | (header_df['ENDDATE'].astype(str).str.strip() == '')]
        
        # Check for invalid date format errors
        invalid_start_dates = header_df[header_df['STARTDATE'].astype(str).str.contains('Format Tanggal Salah', na=False)]
        invalid_end_dates = header_df[header_df['ENDDATE'].astype(str).str.contains('Format Tanggal Salah', na=False)]
        
        if len(empty_start_dates) > 0 or len(empty_end_dates) > 0 or len(invalid_start_dates) > 0 or len(invalid_end_dates) > 0:
            error_details = []
            
            if len(empty_start_dates) > 0:
                error_details.append(f"• Kolom STARTDATE: {len(empty_start_dates)} baris kosong")
            
            if len(empty_end_dates) > 0:
                error_details.append(f"• Kolom ENDDATE: {len(empty_end_dates)} baris kosong")
                
            if len(invalid_start_dates) > 0:
                error_details.append(f"• Kolom STARTDATE: {len(invalid_start_dates)} baris format tanggal salah")
                
            if len(invalid_end_dates) > 0:
                error_details.append(f"• Kolom ENDDATE: {len(invalid_end_dates)} baris format tanggal salah")
            
            error_msg = f"""❌ DATA KOLOM TANGGAL BERMASALAH DI HASIL OUTPUT GIFT

📋 KOLOM YANG MEMILIKI DATA KOSONG DAN MEMBUAT ERROR:
{chr(10).join(error_details)}

💡 SOLUSI:
• Periksa data tanggal di file Excel asli
• Pastikan format tanggal sesuai template (DD/MM/YYYY atau MM/DD/YYYY)
• Contoh format yang benar: 25/12/2025, 12/25/2025
• Download template yang benar dari halaman ini"""
            
            raise ValueError(error_msg)
    
    def validate_date_range(self, df, start_date_col, end_date_col):
        """
        Validate that End_Date is not before Start_Date and check for invalid date formats
        Returns list of invalid rows with details
        """
        invalid_rows = []
        invalid_format_rows = []
        
        for index, row in df.iterrows():
            try:
                start_date_val = row[start_date_col]
                end_date_val = row[end_date_col]

                # Check for empty Start_Date
                if pd.isna(start_date_val) or (isinstance(start_date_val, str) and start_date_val.strip() == ''):
                    invalid_format_rows.append(f"Baris {index + 2}: Kolom '{start_date_col}' kosong")
                
                # Check for empty End_Date
                if pd.isna(end_date_val) or (isinstance(end_date_val, str) and end_date_val.strip() == ''):
                    invalid_format_rows.append(f"Baris {index + 2}: Kolom '{end_date_col}' kosong")

                # If either date was empty, skip further validation for this row
                if any(f"Baris {index + 2}: Kolom '{start_date_col}' kosong" in s for s in invalid_format_rows) or \
                   any(f"Baris {index + 2}: Kolom '{end_date_col}' kosong" in s for s in invalid_format_rows):
                    continue

                # Convert string dates to datetime objects
                start_date = pd.to_datetime(start_date_val, errors='coerce')
                end_date = pd.to_datetime(end_date_val, errors='coerce')
                
                # Check if conversion failed (invalid format)
                if pd.isna(start_date):
                    invalid_format_rows.append(f"Baris {index + 2}: Format tanggal '{start_date_col}' tidak valid ('{start_date_val}')")
                    continue
                    
                if pd.isna(end_date):
                    invalid_format_rows.append(f"Baris {index + 2}: Format tanggal '{end_date_col}' tidak valid ('{end_date_val}')")
                    continue
                
                # Check if end_date is before start_date
                if end_date < start_date:
                    invalid_rows.append({
                        'row': index + 1,  # Excel row number (1-based)
                        'start_date': str(start_date_val),
                        'end_date': str(end_date_val),
                        'sku_bundle': str(row.get('SKU Bundle', 'N/A')),
                        'client': str(row.get('Client', 'N/A')),
                        'marketplace': str(row.get('Marketplace', 'N/A'))
                    })
                    
            except Exception as e:
                self.log(f"Error validating date range at row {index + 1}: {str(e)}", "warning")
                continue
        
        # If there are invalid format errors, raise them first
        if invalid_format_rows:
            error_msg = f"""❌ FORMAT TANGGAL TIDAK VALID

📋 BARIS YANG MEMILIKI FORMAT TANGGAL SALAH:
{chr(10).join(invalid_format_rows)}

💡 SOLUSI:
• Pastikan format tanggal sesuai template (DD/MM/YYYY atau MM/DD/YYYY)
• Contoh format yang benar: 25/12/2025, 12/25/2025
• Hindari format yang salah seperti: 8/32/2025 (bulan tidak boleh > 12)
• Download template yang benar dari halaman ini"""
            raise ValueError(error_msg)
        
        return invalid_rows
    
    def process(self):
        """Main process method that handles the file processing"""
        try:
            self.log(f"Memproses file: {os.path.basename(self.file_path)}")
            
            with pd.ExcelFile(self.file_path) as xls:
                all_sheets_data = {}
                client_name = None
                
                for sheet_name in xls.sheet_names:
                    # Check if sheet is hidden
                    sheet = xls.book[sheet_name]
                    if sheet.sheet_state == 'hidden':
                        self.log(f"Sheet '{sheet_name}' disembunyikan, melewati pemrosesan.")
                        continue
                    
                    self.log(f"Memproses sheet: {sheet_name}")
                    
                    # Read first few rows to determine header rows
                    sample_df = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=2)
                    
                    # Logic to determine header rows
                    header_rows = 0
                    for index, row in sample_df.iterrows():
                        if row.isnull().all():  # If entire row is empty, stop
                            break
                        header_rows += 1
                    
                    # Read DataFrame with determined header rows
                    df = pd.read_excel(xls, sheet_name=sheet_name, header=list(range(header_rows)), dtype=str)
                    
                    # Handle merged cells - Fill empty cells with values from above
                    self.log(f"Menangani cell yang di-merge dengan mengisi nilai kosong...")
                    
                    # Forward fill untuk nilai kosong dari merged cells
                    # Menggunakan metode ffill untuk mengisi nilai yang sama dengan nilai diatasnya
                    for col in df.columns:
                        df[col] = df[col].ffill()
                    
                    # Detect empty cells in rows and fill with values from above rows for the same column
                    df = df.replace('', pd.NA).ffill()
                    
                    # Determine process type based on user selection
                    if self.process_type == "Bundle":
                        processed_df, client_name = self.process_bundle(df)
                    elif self.process_type == "Supplementary":
                        processed_df, client_name = self.process_supplementary(df)
                    elif self.process_type == "Gift":
                        processed_df, client_name = self.process_gift(df)
                    else:
                        raise ValueError(f"Tipe proses tidak valid: {self.process_type}")
                    
                    # Save processed results
                    all_sheets_data[sheet_name] = processed_df
                
                if self.combine_sheets and len(all_sheets_data) > 1:
                    # Combine all sheets if option is selected
                    if isinstance(next(iter(all_sheets_data.values())), dict):
                        # For Gift type with multiple dataframes
                        combined_header = pd.concat([sheet_data["Header"] for sheet_data in all_sheets_data.values()], ignore_index=True)
                        combined_line = pd.concat([sheet_data["Line"] for sheet_data in all_sheets_data.values()], ignore_index=True)
                        all_sheets_data = {"Combined": {"Header": combined_header, "Line": combined_line}}
                    else:
                        # For other types with single dataframe
                        combined_df = pd.concat(all_sheets_data.values(), ignore_index=True)
                        all_sheets_data = {"Combined": combined_df}
                
                # Save results to file with client name
                output_path = self.save_output(all_sheets_data, client_name)
                self.output_path = output_path
                
                # Prepare analytics data
                analytics_data = self.prepare_analytics_data(all_sheets_data, client_name)
                
                return {
                    "status": "success", 
                    "logs": self.logs, 
                    "output_path": output_path,
                    "analytics_data": analytics_data
                }
                
        except pd.errors.EmptyDataError:
            error_msg = "❌ FILE EXCEL KOSONG\n\nFile Excel yang diunggah kosong atau tidak berisi data yang valid.\n\n💡 SOLUSI:\n• Download dan gunakan template yang benar"
            self.log(error_msg, "error")
            return {"status": "error", "logs": self.logs, "error": error_msg}
        except pd.errors.ParserError as e:
            error_msg = f"❌ FILE EXCEL RUSAK\n\nFile Excel rusak atau tidak dapat dibaca.\nError: {str(e)}\n\n💡 SOLUSI:\n• Coba buka file di Excel dan simpan ulang"
            self.log(error_msg, "error")
            return {"status": "error", "logs": self.logs, "error": error_msg}
        except ValueError as e:
            error_msg = f"❌ DATA TIDAK VALID\n\n{str(e)}\n\n💡 SOLUSI:\n• Periksa format data sesuai dengan template yang diharapkan\n• Pastikan semua kolom wajib terisi dengan benar\n• Download template yang benar dari halaman ini"
            self.log(error_msg, "error")
            return {"status": "error", "logs": self.logs, "error": error_msg}
        except Exception as e:
            error_msg = self.generate_detailed_error_message(e)
            self.log(error_msg, "error")
            logging.error(f"Error processing file: {str(e)}")
            return {"status": "error", "logs": self.logs, "error": error_msg}
    
    def process_bundle(self, df):
        """Process bundle data"""
        try:
            # Kolom yang diperlukan
            kolom_wajib = ["SKU Bundle", "SKU Product", "Qty", "Client", "Marketplace", "Start_Date", "End_Date"]
            
            # Debug: Cetak nama kolom sebelum pemrosesan
            self.log("Kolom asli: " + str(df.columns.tolist()))
            
            # Handle merged cells - Additional check for empty strings and spaces
            df = df.replace(['', ' ', None], pd.NA)
            
            # Handle merged cells in each column independently
            for col in df.columns:
                df[col] = df[col].ffill()
            
            # Combine headers and create appropriate column names
            df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
            
            # Debug: Cetak nama kolom setelah penggabungan
            self.log("Kolom setelah penggabungan: " + str(df.columns.tolist()))
            
            # Validate required columns
            missing_columns = []
            for kolom in kolom_wajib:
                if not any(kolom.lower() in col.lower() for col in df.columns):
                    missing_columns.append(kolom)
            
            if missing_columns:
                error_msg = f"❌ KOLOM WAJIB TIDAK DITEMUKAN\n\nKolom berikut tidak ditemukan dalam file Excel:\n"
                for i, kolom in enumerate(missing_columns, 1):
                    error_msg += f"{i}. {kolom}\n"
                error_msg += f"\n💡 SOLUSI:\n"
                error_msg += f"• Pastikan file Excel memiliki kolom-kolom tersebut\n"
                error_msg += f"• Periksa nama kolom (case sensitive)\n"
                error_msg += f"• Download template yang benar dari halaman ini\n"
                error_msg += f"• Pastikan tidak ada spasi ekstra di nama kolom"
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            # Column mapping for standardization
            column_mapping = {
                # ... existing code ...
                
                # Additional form-specific mappings
                'SKU BUNDLING SKU Bundle': 'SKU Bundle',
                'SKU BUNDLING Name Bundle': 'Name Bundle',
                'SKU Komponen Name Product': 'Name Product',
                'Form Bundle : Client': 'Client',
                'Form Bundle : End_Date': 'End_Date',
                'Form Bundle : Marketplace': 'Marketplace',
                'Form Bundle : SKU BUNDLE': 'SKU Bundle',
                'Form Bundle : SKU Komponen': 'SKU Product',
                'Form Bundle : SKU Komponen.2': 'Qty',
                'Form Bundle : SKU Product': 'SKU Product',
                'Form Bundle : Start_Date': 'Start_Date',
                'Form Bundling List : Client': 'Client',
                'Form Bundling List : End_Date': 'End_Date',
                'Form Bundling List : Marketplace': 'Marketplace',
                'Form Bundling List : SKU Bundling List': 'SKU Bundle',
                'Form Bundling List : SKU Komponen': 'SKU Product',
                'Form Bundling List : SKU Komponen.2': 'Qty',
                'Form Bundling List : Start_Date': 'Start_Date',
                'SKU BUNDLE SKU Bundle': 'SKU Bundle',
                'SKU BUNDLE Name Bundle': 'Name Bundle',
                'SKU Bundling List SKU Bundle': 'SKU Bundle',
                'SKU Komponen Qty': 'Qty',
                'SKU Komponen SKU Product': 'SKU Product',
                
                # Client mappings 
                'Client Unnamed: 0_level_1': 'Client',
                'Client Unnamed: 1_level_1': 'Client',
                'Client Unnamed: 2_level_1': 'Client',
                'Client Unnamed: 3_level_1': 'Client',
                'Client Unnamed: 4_level_1': 'Client',
                'Client Unnamed: 5_level_1': 'Client',
                'Client Unnamed: 6_level_1': 'Client',
                'Client Unnamed: 7_level_1': 'Client',
                'Client Unnamed: 8_level_1': 'Client',
                'Client Unnamed: 9_level_1': 'Client',
                'Client Unnamed: 10_level_1': 'Client',
                'Client Unnamed: 11_level_1': 'Client',
                'Client Unnamed: 12_level_1': 'Client',
                'Client Unnamed: 13_level_1': 'Client',
                'Client Unnamed: 14_level_1': 'Client',
                'Client Unnamed: 15_level_1': 'Client',
                'Client Unnamed: 16_level_1': 'Client',
                'Client Unnamed: 17_level_1': 'Client',
                'Client Unnamed: 18_level_1': 'Client',
                'Client Unnamed: 19_level_1': 'Client',
                'Client Unnamed: 20_level_1': 'Client',
                
                # Marketplace mappings
                'Marketplace Unnamed: 0_level_1': 'Marketplace',
                'Marketplace Unnamed: 1_level_1': 'Marketplace',
                'Marketplace Unnamed: 2_level_1': 'Marketplace',
                'Marketplace Unnamed: 3_level_1': 'Marketplace',
                'Marketplace Unnamed: 4_level_1': 'Marketplace',
                'Marketplace Unnamed: 5_level_1': 'Marketplace',
                'Marketplace Unnamed: 6_level_1': 'Marketplace',
                'Marketplace Unnamed: 7_level_1': 'Marketplace',
                'Marketplace Unnamed: 8_level_1': 'Marketplace',
                'Marketplace Unnamed: 9_level_1': 'Marketplace',
                'Marketplace Unnamed: 10_level_1': 'Marketplace',
                'Marketplace Unnamed: 11_level_1': 'Marketplace',
                'Marketplace Unnamed: 12_level_1': 'Marketplace',
                'Marketplace Unnamed: 13_level_1': 'Marketplace',
                'Marketplace Unnamed: 14_level_1': 'Marketplace',
                'Marketplace Unnamed: 15_level_1': 'Marketplace',
                'Marketplace Unnamed: 16_level_1': 'Marketplace',
                'Marketplace Unnamed: 17_level_1': 'Marketplace',
                'Marketplace Unnamed: 18_level_1': 'Marketplace',
                'Marketplace Unnamed: 19_level_1': 'Marketplace',
                'Marketplace Unnamed: 20_level_1': 'Marketplace',
                'Market Place Unnamed: 0_level_1': 'Marketplace',
                'Market Place Unnamed: 1_level_1': 'Marketplace',
                'Market Place Unnamed: 2_level_1': 'Marketplace',
                'Market Place Unnamed: 3_level_1': 'Marketplace',
                'Market Place Unnamed: 4_level_1': 'Marketplace',
                'Market Place Unnamed: 5_level_1': 'Marketplace',
                'Market Place Unnamed: 6_level_1': 'Marketplace',
                'Market Place Unnamed: 7_level_1': 'Marketplace',
                'Market Place Unnamed: 8_level_1': 'Marketplace',
                'Market Place Unnamed: 9_level_1': 'Marketplace',
                'Market Place Unnamed: 10_level_1': 'Marketplace',
                'Market Place Unnamed: 11_level_1': 'Marketplace',
                'Market Place Unnamed: 12_level_1': 'Marketplace',
                'Market Place Unnamed: 13_level_1': 'Marketplace',
                'Market Place Unnamed: 14_level_1': 'Marketplace',
                'Market Place Unnamed: 15_level_1': 'Marketplace',
                'Market Place Unnamed: 16_level_1': 'Marketplace',
                'Market Place Unnamed: 17_level_1': 'Marketplace',
                'Market Place Unnamed: 18_level_1': 'Marketplace',
                'Market Place Unnamed: 19_level_1': 'Marketplace',
                'Market Place Unnamed: 20_level_1': 'Marketplace',
                
                # Start_Date mappings
                'Start_Date Unnamed: 0_level_1': 'Start_Date',
                'Start_Date Unnamed: 1_level_1': 'Start_Date',
                'Start_Date Unnamed: 2_level_1': 'Start_Date',
                'Start_Date Unnamed: 3_level_1': 'Start_Date',
                'Start_Date Unnamed: 4_level_1': 'Start_Date',
                'Start_Date Unnamed: 5_level_1': 'Start_Date',
                'Start_Date Unnamed: 6_level_1': 'Start_Date',
                'Start_Date Unnamed: 7_level_1': 'Start_Date',
                'Start_Date Unnamed: 8_level_1': 'Start_Date',
                'Start_Date Unnamed: 9_level_1': 'Start_Date',
                'Start_Date Unnamed: 10_level_1': 'Start_Date',
                'Start_Date Unnamed: 11_level_1': 'Start_Date',
                'Start_Date Unnamed: 12_level_1': 'Start_Date',
                'Start_Date Unnamed: 13_level_1': 'Start_Date',
                'Start_Date Unnamed: 14_level_1': 'Start_Date',
                'Start_Date Unnamed: 15_level_1': 'Start_Date',
                'Start_Date Unnamed: 16_level_1': 'Start_Date',
                'Start_Date Unnamed: 17_level_1': 'Start_Date',
                'Start_Date Unnamed: 18_level_1': 'Start_Date',
                'Start_Date Unnamed: 19_level_1': 'Start_Date',
                'Start_Date Unnamed: 20_level_1': 'Start_Date',
                'Periode Start_Date': 'Start_Date',
                'Tanggal Mulai': 'Start_Date',
                'Tgl Mulai': 'Start_Date',
                
                # End_Date mappings
                'End_Date Unnamed: 0_level_1': 'End_Date',
                'End_Date Unnamed: 1_level_1': 'End_Date',
                'End_Date Unnamed: 2_level_1': 'End_Date',
                'End_Date Unnamed: 3_level_1': 'End_Date',
                'End_Date Unnamed: 4_level_1': 'End_Date',
                'End_Date Unnamed: 5_level_1': 'End_Date',
                'End_Date Unnamed: 6_level_1': 'End_Date',
                'End_Date Unnamed: 7_level_1': 'End_Date',
                'End_Date Unnamed: 8_level_1': 'End_Date',
                'End_Date Unnamed: 9_level_1': 'End_Date',
                'End_Date Unnamed: 10_level_1': 'End_Date',
                'End_Date Unnamed: 11_level_1': 'End_Date',
                'End_Date Unnamed: 12_level_1': 'End_Date',
                'End_Date Unnamed: 13_level_1': 'End_Date',
                'End_Date Unnamed: 14_level_1': 'End_Date',
                'End_Date Unnamed: 15_level_1': 'End_Date',
                'End_Date Unnamed: 16_level_1': 'End_Date',
                'End_Date Unnamed: 17_level_1': 'End_Date',
                'End_Date Unnamed: 18_level_1': 'End_Date',
                'End_Date Unnamed: 19_level_1': 'End_Date',
                'End_Date Unnamed: 20_level_1': 'End_Date',
                'Periode End_Date': 'End_Date',
                'Tanggal Selesai': 'End_Date',
                'Tgl Selesai': 'End_Date',
                
                # SKU Bundle mappings
                'SKU Bundle Unnamed: 0_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 1_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 2_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 3_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 4_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 5_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 6_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 7_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 8_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 9_level_1': 'SKU Bundle',
                'SKU Bundle Unnamed: 10_level_1': 'SKU Bundle',
                
                # SKU Product mappings
                'SKU Product Unnamed: 0_level_1': 'SKU Product',
                'SKU Product Unnamed: 1_level_1': 'SKU Product',
                'SKU Product Unnamed: 2_level_1': 'SKU Product',
                'SKU Product Unnamed: 3_level_1': 'SKU Product',
                'SKU Product Unnamed: 4_level_1': 'SKU Product',
                'SKU Product Unnamed: 5_level_1': 'SKU Product',
                'SKU Product Unnamed: 6_level_1': 'SKU Product',
                'SKU Product Unnamed: 7_level_1': 'SKU Product',
                'SKU Product Unnamed: 8_level_1': 'SKU Product',
                'SKU Product Unnamed: 9_level_1': 'SKU Product',
                'SKU Product Unnamed: 10_level_1': 'SKU Product',
                
                # Qty mappings
                'Qty Unnamed: 0_level_1': 'Qty',
                'Qty Unnamed: 1_level_1': 'Qty',
                'Qty Unnamed: 2_level_1': 'Qty',
                'Qty Unnamed: 3_level_1': 'Qty',
                'Qty Unnamed: 4_level_1': 'Qty',
                'Qty Unnamed: 5_level_1': 'Qty',
                'Qty Unnamed: 6_level_1': 'Qty',
                'Qty Unnamed: 7_level_1': 'Qty',
                'Qty Unnamed: 8_level_1': 'Qty',
                'Qty Unnamed: 9_level_1': 'Qty',
                'Qty Unnamed: 10_level_1': 'Qty',
                'Quantity Unnamed: 0_level_1': 'Qty',
                'Quantity Unnamed: 1_level_1': 'Qty',
                'Quantity Unnamed: 2_level_1': 'Qty',
                'Quantity Unnamed: 3_level_1': 'Qty',
                'Quantity Unnamed: 4_level_1': 'Qty',
                'Quantity Unnamed: 5_level_1': 'Qty',
                'Quantity Unnamed: 6_level_1': 'Qty',
                'Quantity Unnamed: 7_level_1': 'Qty',
                'Quantity Unnamed: 8_level_1': 'Qty',
                'Quantity Unnamed: 9_level_1': 'Qty',
                'Quantity Unnamed: 10_level_1': 'Qty',
                'Jumlah': 'Qty',
                'Jumlah Unnamed: 0_level_1': 'Qty',
                'Jumlah Unnamed: 1_level_1': 'Qty',
                
                # HTTP headers (not actually relevant to Excel processing)
                'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36': 'Sec-Ch-Ua',
                'Sec-Ch-Ua': 'Google Chrome',
                'Content-Type': 'multipart/form-data',
                'Sec-Ch-Ua-Mobile': '0',
                'Accept': '*/*',
                'Origin': 'http://localhost:3000',
                'Sec-Fetch-Site': 'cross-site',
                'Sec-Fetch-Mode': 'cors',
                'Sec-Fetch-Dest': 'empty',
                'Referer': 'http://localhost:3000/',
                'Accept-Encoding': 'gzip, deflate, br, zstd',
                'Accept-Language': 'en-US,en',
            }
            
            # Membuat dictionary baru untuk case-insensitive mapping
            lowercase_column_mapping = {}
            for key, value in column_mapping.items():
                lowercase_column_mapping[key.lower()] = value
            
            # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
            rename_dict = {}
            for col in df.columns:
                col_lower = col.lower()
                if col_lower in lowercase_column_mapping:
                    rename_dict[col] = lowercase_column_mapping[col_lower]
            
            # Rename kolom menggunakan dictionary yang baru dibuat
            df = df.rename(columns=rename_dict)
            df = df.copy()
            df.ffill(inplace=True)
            
            # Debug: Cetak nama kolom setelah mapping
            self.log("Kolom setelah mapping: " + str(df.columns.tolist()))
            
            # Debug: tambahan untuk membantu diagnosa
            self.log("Format exact dari kolom input (utk debug): " + str([col for col in df.columns if 'market' in col.lower() or 'place' in col.lower()]))
            
            # Cek apakah kolom yang diperlukan ada
            for kolom in kolom_wajib:
                if kolom not in df.columns:
                    raise KeyError(f"Kolom '{kolom}' tidak ditemukan")
            
            # Bersihkan spasi dan baris baru pada kolom SKU Bundle dan SKU Product
            df["SKU Bundle"] = df["SKU Bundle"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True)
            df["SKU Product"] = df["SKU Product"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True)
            
            # Validate date ranges before processing
            self.log("Memvalidasi range tanggal...")
            invalid_date_rows = self.validate_date_range(df, 'Start_Date', 'End_Date')
            
            if invalid_date_rows:
                # Create detailed error message
                error_details = []
                for row_info in invalid_date_rows:
                    error_details.append(
                        f"Baris {row_info['row']}: SKU '{row_info['sku_bundle']}' "
                        f"(Client: {row_info['client']}, Marketplace: {row_info['marketplace']}) - "
                        f"Start_Date: {row_info['start_date']}, End_Date: {row_info['end_date']}"
                    )
                
                error_msg = (
                    f"❌ TANGGAL TIDAK VALID\n\n"
                    f"Ditemukan {len(invalid_date_rows)} baris dengan tanggal yang tidak valid.\n"
                    f"End_Date tidak boleh lebih awal dari Start_Date.\n\n"
                    f"📋 DETAIL ERROR:\n" + "\n".join(error_details) + "\n\n"
                    f"💡 SOLUSI:\n"
                    f"• Periksa dan perbaiki tanggal di baris-baris tersebut\n"
                    f"• Pastikan End_Date selalu setelah Start_Date\n"
                    f"• Format tanggal harus konsisten (DD/MM/YYYY atau YYYY-MM-DD)"
                )
                
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            self.log("Validasi tanggal berhasil - semua range tanggal valid")
            
            # Deteksi kolom Name Product
            name_product_column = None
            name_product_candidates = [
                "Name Product", "Product Name", "Nama Produk", "Nama", "Name", 
                "Nama Item", "Item Name", "Deskripsi Produk", "Product Description", 
                "Description", "Deskripsi", "SKU Name", "SKU Nama"
            ]
            
            # Debug: Cetak nama kolom yang tersedia
            self.log("Mencari kolom Name Product dari kolom: " + str(df.columns.tolist()))
            
            # Cari kolom yang berisi nama produk
            for candidate in name_product_candidates:
                if candidate in df.columns:
                    name_product_column = candidate
                    self.log(f"Kolom name_product ditemukan: {name_product_column}")
                    break
            
            # Jika tidak ada yang cocok, coba cari kolom dengan nama yang mirip
            if name_product_column is None:
                # Cari kolom yang mengandung kata-kata tertentu
                for col in df.columns:
                    col_lower = col.lower()
                    # Cari kolom yang sepertinya berisi nama produk tapi bukan nama bundle
                    if (('name' in col_lower or 'nama' in col_lower) and 
                        ('product' in col_lower or 'produk' in col_lower or 'item' in col_lower) and 
                        'bundle' not in col_lower):
                        name_product_column = col
                        self.log(f"Kolom name_product mirip ditemukan: {name_product_column}")
                        break
                
                # Jika masih tidak ditemukan, gunakan heuristik tambahan
                if name_product_column is None:
                    # Periksa apakah ada kolom dengan "Name" atau "Nama" di dalamnya yang bukan bundle
                    for col in df.columns:
                        col_lower = col.lower()
                        if ('name' in col_lower or 'nama' in col_lower) and 'bundle' not in col_lower:
                            name_product_column = col
                            self.log(f"Kolom name_product berdasarkan heuristik ditemukan: {name_product_column}")
                            break
            
            # Validasi bahwa satu SKU Product hanya memiliki satu Name Product
            if name_product_column is not None:
                self.log(f"Melakukan validasi Name Product menggunakan kolom: {name_product_column}")
                
                # Buat dictionary untuk memetakan SKU Product ke semua Name Product
                sku_to_names = {}
                
                # Iterasi melalui baris data
                for idx, row in df.iterrows():
                    # Ambil SKU dan Name Product
                    sku = str(row["SKU Product"]).strip() if pd.notna(row["SKU Product"]) else None
                    name = str(row[name_product_column]).strip() if pd.notna(row[name_product_column]) else None
                    
                    # Jika keduanya valid
                    if sku and name:
                        # Tambahkan ke dictionary
                        if sku not in sku_to_names:
                            sku_to_names[sku] = set()
                        sku_to_names[sku].add(name)
                
                # Temukan SKU dengan lebih dari satu nama
                problematic_skus = {sku: names for sku, names in sku_to_names.items() if len(names) > 1}
                
                # Jika ada SKU bermasalah, tampilkan pesan error
                if problematic_skus:
                    error_msg = f"❌ SKU PRODUCT DUPLIKAT\n\n"
                    error_msg += f"Ditemukan {len(problematic_skus)} SKU Product dengan lebih dari satu Name Product.\n"
                    error_msg += f"Satu SKU Product hanya boleh memiliki satu Name Product.\n\n"
                    
                    # Tambahkan daftar SEMUA SKU yang bermasalah tanpa batasan
                    error_msg += "📋 DAFTAR SKU YANG BERMASALAH:\n"
                    
                    # Tampilkan semua SKU yang bermasalah
                    for idx, (sku, names) in enumerate(problematic_skus.items(), 1):
                        error_msg += f"{idx}. SKU: {sku}\n"
                        names_list = list(names)
                        for name in names_list:
                            error_msg += f"   • Name Product: {name}\n"
                        error_msg += "\n"
                    
                    error_msg += "💡 SOLUSI:\n"
                    error_msg += "• Pastikan setiap SKU Product hanya memiliki satu Name Product\n"
                    error_msg += "• Periksa dan perbaiki data yang duplikat\n"
                    error_msg += "• Hapus atau gabungkan baris yang memiliki SKU sama"
                    raise ValueError(error_msg)
                
                self.log(f"Validasi SKU Product dengan Name Product: {len(sku_to_names)} SKU diperiksa, semua valid.")
            
            # Validasi data kosong sebelum pemrosesan
            self.validate_required_columns_data(df)
            
            # Hapus baris yang memiliki nilai kosong pada kolom BOMSKU
            df = df[df["SKU Product"].notna() & (df["SKU Product"] != "")]
            
            # Validasi: SKU Bundle tidak boleh sama dengan SKU Product
            sku_conflicts = df[df["SKU Bundle"] == df["SKU Product"]]
            if not sku_conflicts.empty:
                # Format pesan error seperti contoh yang diminta
                error_msg = f"Error: Kesalahan saat memproses file, ditemukan {len(sku_conflicts)} baris di mana SKU Bundle sama dengan SKU Product.\n\n"
                
                # Tambahkan daftar SEMUA SKU yang bermasalah tanpa batasan
                error_msg += "Daftar SKU yang bermasalah:\n"
                
                # Tampilkan semua SKU yang bermasalah
                for idx, (i, row) in enumerate(sku_conflicts.iterrows(), 1):
                    error_msg += f"{idx}. {row['SKU Bundle']}\n"
                
                error_msg += "\nProses dihentikan. Mohon perbaiki data agar SKU Bundle tidak sama dengan SKU Product."
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            self.log("Validasi SKU Bundle tidak sama dengan SKU Product: Semua data valid.")
            
            # Validasi: SKU Bundle tidak boleh muncul sebagai SKU Product di baris lain
            # Ambil list semua SKU Bundle dan SKU Product
            all_sku_bundles = set(df["SKU Bundle"].unique())
            all_sku_products = set(df["SKU Product"].unique())
            
            # Cari SKU Bundle yang juga muncul sebagai SKU Product
            bundle_as_product = all_sku_bundles.intersection(all_sku_products)
            
            if bundle_as_product:
                # Identifikasi sheet dan baris untuk SKU yang bermasalah
                problematic_rows = []
                for sku in bundle_as_product:
                    # Temukan baris di mana SKU tersebut muncul sebagai SKU Product
                    rows = df[df["SKU Product"] == sku].index.tolist()
                    for row in rows:
                        sheet_name = getattr(df, 'sheet_name', 'data')  # Gunakan sheet_name jika ada, atau 'data' sebagai default
                        row_number = row + 2  # +2 karena biasanya header Excel ada di baris 1 dan indeks pandas dimulai dari 0
                        problematic_rows.append((sheet_name, row_number, sku))
                
                # Buat pesan error
                if problematic_rows:
                    # Format pesan error
                    error_msg = f"Error: Ditemukan {len(bundle_as_product)} SKU yang digunakan sebagai SKU Bundle dan juga muncul sebagai SKU Product.\n\n"
                    
                    # Tambahkan daftar SKU yang bermasalah
                    error_msg += "Daftar SKU yang bermasalah:\n"
                    for idx, sku in enumerate(bundle_as_product, 1):
                        error_msg += f"{idx}. {sku}\n"
                    
                    # Contoh detail baris yang bermasalah
                    first_issue = problematic_rows[0]
                    error_msg += f"\nContoh: Di sheet '{first_issue[0]}', baris ke-{first_issue[1]} ditemukan SKU {first_issue[2]} yang digunakan sebagai SKU Bundle dan juga muncul sebagai SKU Product."
                    error_msg += "\n\nProses dihentikan. Mohon perbaiki data agar nilai SKU Bundle tidak muncul di kolom SKU Product."
                    
                    self.log(error_msg, "error")
                    raise ValueError(error_msg)
            
            self.log("Validasi SKU Bundle tidak muncul sebagai SKU Product di baris lain: Semua data valid.")
            
            # Validasi: SKU Product harus unik dalam satu grup SKU Bundle
            # Cek duplikasi SKU Product dalam grup SKU Bundle yang sama
            duplicate_products = df.groupby('SKU Bundle')['SKU Product'].apply(lambda x: x[x.duplicated()]).reset_index()
            
            if not duplicate_products.empty:
                # Format pesan error
                error_msg = f"Error: Kesalahan saat memproses file, ditemukan {len(duplicate_products)} baris di mana SKU Product duplikat dalam grup SKU Bundle yang sama.\n\n"
                
                # Tambahkan daftar konflik duplikasi
                error_msg += "Daftar duplikasi yang ditemukan:\n"
                
                # Ambil uniqueww nilai SKU Bundle yang memiliki duplikasi
                problematic_bundles = duplicate_products['SKU Bundle'].unique()
                
                # Tampilkan informasi untuk setiap SKU Bundle bermasalah
                for idx, bundle_sku in enumerate(problematic_bundles, 1):
                    # Filter baris dengan SKU Bundle tertentu
                    bundle_rows = df[df['SKU Bundle'] == bundle_sku]
                    # Temukan duplikasi SKU Product
                    duplicate_products_in_bundle = bundle_rows['SKU Product'][bundle_rows['SKU Product'].duplicated(keep=False)]
                    # Ambil nilai unik SKU Product yang duplikat
                    unique_duplicate_products = duplicate_products_in_bundle.unique()
                    # Tambahkan ke pesan error
                    error_msg += f"{idx}. SKU Bundle: {bundle_sku}\n"
                    for prod_idx, product_sku in enumerate(unique_duplicate_products, 1):
                        error_msg += f"   - SKU Product duplikat: {product_sku}\n"
                
                error_msg += "\nProses dihentikan. Mohon perbaiki data agar SKU Product tidak duplikat dalam grup SKU Bundle yang sama."
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            self.log("Validasi SKU Product tidak duplikat dalam grup SKU Bundle: Semua data valid.")
            
            # Tampilkan pesan sukses validasi
            self.log("Semua data valid. Tidak ada konflik atau duplikasi.")
            
            # Pisahkan nilai dalam kolom Marketplace dan gandakan baris
            if 'Marketplace' in df.columns:
                df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
                df['Marketplace'] = df['Marketplace'].str.strip()
            else:
                raise KeyError("Kolom 'Marketplace' tidak ditemukan setelah pemetaan.")
            
            # Assign ShopId dan validasi (setelah explode untuk multiple marketplace)
            df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
            
            # Validasi ShopId yang menghasilkan "Unknown"
            self.validate_shop_id_mapping(df)
            
            # Konversi format tanggal
            if 'Start_Date' in df.columns:
                df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
            else:
                self.log("Kolom yang tersedia: " + str(df.columns.tolist()))
                raise KeyError("Kolom 'Start_Date' tidak ditemukan")
                
            if 'End_Date' in df.columns:
                df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
            else:
                raise KeyError("Kolom 'End_Date' tidak ditemukan")
            
            # Assign ShopId sudah dilakukan di atas
            
            # Simpan urutan asli dari file input
            df['original_index'] = df.index
            
            # Urutkan berdasarkan Marketplace agar hasil dikelompokkan per marketplace
            # Hal ini hanya dilakukan untuk tipe proses Bundle
            if self.process_type == "Bundle":
                self.log("Mengurutkan data berdasarkan Marketplace")
                
                # Mendefinisikan urutan prioritas marketplace (opsional)
                marketplace_priority = {
                    'shopee': 1,
                    'tiktok': 2,
                    'tokopedia': 3,
                    'lazada': 4,
                    'blibli': 5,
                    'zalora': 6,
                    'desty': 7,
                    'jubelio': 8
                }
                
                # Tambahkan kolom prioritas untuk pengurutan
                df['marketplace_priority'] = df['Marketplace'].str.lower().map(
                    lambda x: next((priority for mkt, priority in marketplace_priority.items() 
                                  if mkt in x.lower()), 999)  # Default 999 jika tidak cocok
                )
                
                # Urutkan berdasarkan prioritas marketplace, lalu berdasarkan nama marketplace
                df = df.sort_values(by=['marketplace_priority', 'Marketplace'])
                
                # Hapus kolom prioritas setelah pengurutan
                df = df.drop(columns=['marketplace_priority'])
            
            # Hapus kolom Client dari output
            result_df = pd.DataFrame({
                "MainSKU": df["SKU Bundle"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True),
                "BOMSKU": df["SKU Product"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True),
                "BOMQty": df["Qty"],
                "IsActive": True,
                "ShopId": df["ShopId"],
                "CreatedDate": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
                "CreatedBy": self.created_by,
                "UpdatedDate": "NULL",
                "UpdatedBy": "NULL",
                "SingleSKU": False,
                "IsHadiah": False,
                "StartDate": df["Start_Date"],
                "EndDate": df["End_Date"],
                "Multiply": True
            })
            
            # Validasi hasil output untuk kolom tanggal
            self.validate_output_date_columns(result_df)
            
            # Mapping untuk marketplace_id
            marketplace_id_mapping = {
                'tokopedia': 1,
                'shopee': 2,
                'lazada': 3,
                'zalora': 6,
                'blibli': 7,
                'jdid': 8,
                'jubelio': 9,
                'shopify': 10,
                'tiktok': 11,
                'desty': 23
            }
            
            # Fungsi untuk mendapatkan marketplace_id berdasarkan nama marketplace
            def get_marketplace_id(marketplace_name):
                if not marketplace_name or pd.isna(marketplace_name):
                    return None
                
                try:
                    marketplace_name = str(marketplace_name).lower().strip()
                except AttributeError as e:
                    self.log(f"Error converting marketplace name to string: {marketplace_name}, type: {type(marketplace_name)}", "warning")
                    return None
                
                # Cek exact match di mapping
                if marketplace_name in marketplace_id_mapping:
                    return marketplace_id_mapping[marketplace_name]
                
                # Cek partial match (jika marketplace_name mengandung key di mapping)
                for key, value in marketplace_id_mapping.items():
                    if key in marketplace_name or marketplace_name in key:
                        return value
                
                # Jika tidak ada match, coba match secara lebih fleksibel
                # Misalnya "tokped" untuk "tokopedia", "sp" untuk "shopee", dll.
                common_aliases = {
                    'tokped': 1,
                    'tp': 1,
                    'sp': 2,
                    'lzd': 3,
                    'zlr': 6,
                    'bb': 7,
                    'jd': 8,
                    'jub': 9,
                    'sf': 10,
                    'tt': 11,
                    'dst': 23
                }
                
                for alias, id_value in common_aliases.items():
                    if alias in marketplace_name:
                        return id_value
                
                # Default return None jika tidak ada match
                return None
            
            # Mendeteksi kolom name bundle
            name_bundle_column = None
            name_bundle_candidates = ["SKU BUNDLE Name Bundle", "Name Bundle", "Product Name", "Item Name", "Bundle Name", "Name", "Nama Produk", "Nama Bundle"]
            
            # Debug: Cetak nama kolom yang tersedia
            self.log("Kolom yang tersedia: " + str(df.columns.tolist()))
            
            # Periksa semua kandidat nama kolom
            for candidate in name_bundle_candidates:
                if candidate in df.columns:
                    name_bundle_column = candidate
                    self.log(f"Kolom name_bundle ditemukan: {name_bundle_column}")
                    break
            
            # Jika tidak ada yang cocok, coba cari kolom dengan nama yang mirip
            if name_bundle_column is None:
                for col in df.columns:
                    col_lower = col.lower()
                    if 'name' in col_lower or 'nama' in col_lower or 'product' in col_lower or 'produk' in col_lower:
                        name_bundle_column = col
                        self.log(f"Kolom name_bundle mirip ditemukan: {name_bundle_column}")
                        break
            
            # Gunakan nilai dari kolom name_bundle_column jika ditemukan, jika tidak gunakan default "Name Bundle"
            if name_bundle_column is not None:
                # Debug: Cetak nilai dari kolom name_bundle
                self.log(f"Nilai dari kolom {name_bundle_column}: " + str(df[name_bundle_column].head().tolist()))
                # Gunakan metode tolist atau values untuk mendapatkan array numpy
                product_name_values = df[name_bundle_column].values
            else:
                # Gunakan nilai default yang lebih cocok
                self.log("Tidak ada kolom name_bundle ditemukan, menggunakan nilai default")
                product_name_values = ["Bundle Product"] * len(df)
            
            # Debug final
            self.log("Tipe product_name_values: " + str(type(product_name_values)))
            if hasattr(product_name_values, 'shape'):
                self.log("Bentuk product_name_values: " + str(product_name_values.shape))
            
            # Buat DataFrame cms_product_df
            cms_product_df = pd.DataFrame({
                "product_id": None,
                "client_id": df["ShopId"].apply(get_client_id),
                "marketplace_id": df["Marketplace"].apply(get_marketplace_id), 
                "shop_id": df["ShopId"],
                "product_code": df["SKU Bundle"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True),
                "product_name": product_name_values,  # Gunakan array values
                "product_desc": product_name_values,  # Gunakan array values
                "barcode": df["SKU Bundle"].astype(str).str.strip().str.replace(r'[\r\n\t]+', '', regex=True),
                "product_source": "NULL",
                "product_type_id": "2",
                "product_category_id": "7",
                "product_uom_id": "1",
                "product_cogst": "0",
                "sales_price": "0",
                "product_discount": "0",
                "parent_product": "NULL",
                "status": "TRUE",
                "created_at": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
                "created_by": self.created_by,
                "updated_at": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
                "updated_by": self.created_by,
                "remark": "NULL",
                "tier": "TRUE"
            })
            
            # Menghilangkan duplikasi berdasarkan product_code (SKU Bundle)
            self.log(f"Jumlah baris cms_product sebelum deduplikasi: {len(cms_product_df)}")
            
            # Periksa apakah ada duplikasi berdasarkan kombinasi product_code dan marketplace_id
            cms_product_df['product_marketplace_combined'] = cms_product_df['product_code'].astype(str) + '_' + cms_product_df['marketplace_id'].astype(str)
            
            # Periksa apakah ada duplikasi pada kombinasi
            has_duplicates = cms_product_df["product_marketplace_combined"].duplicated().any()
            
            if has_duplicates:
                # Tampilkan informasi tentang duplikasi
                dup_counts = cms_product_df["product_marketplace_combined"].value_counts()
                self.log(f"Ditemukan kombinasi product_code dan marketplace_id dengan duplikasi:")
                for code, count in dup_counts[dup_counts > 1].items():
                    self.log(f"  - {code}: {count} kemunculan")
                
                # Hapus duplikasi berdasarkan kombinasi, bukan hanya product_code
                cms_product_df = cms_product_df.drop_duplicates(subset=["product_marketplace_combined"], keep="first")
                self.log(f"Duplikasi pada kombinasi product_code dan marketplace_id telah dihapus.")
            else:
                # Periksa duplikasi berdasarkan product_code saja (untuk backward compatibility)
                has_prod_duplicates = cms_product_df["product_code"].duplicated().any()
                if has_prod_duplicates:
                    self.log("Tidak ada duplikasi pada kombinasi, tetapi ada pada product_code saja.")
                    self.log("Membiarkan product_code yang sama dengan marketplace berbeda.")
                    
                    # Tampilkan contoh product_code yang muncul di beberapa marketplace
                    dup_product_codes = cms_product_df[cms_product_df.duplicated(subset=["product_code"], keep=False)]
                    if not dup_product_codes.empty:
                        example_code = dup_product_codes["product_code"].iloc[0]
                        example_rows = cms_product_df[cms_product_df["product_code"] == example_code]
                        self.log(f"\nContoh product_code '{example_code}' muncul di marketplace:")
                        for idx, row in example_rows.iterrows():
                            marketplace_id = row["marketplace_id"]
                            self.log(f"  - Marketplace ID: {marketplace_id}")
            
            # Hapus kolom bantuan
            if 'product_marketplace_combined' in cms_product_df.columns:
                cms_product_df = cms_product_df.drop(columns=['product_marketplace_combined'])
            
            self.log(f"Jumlah baris cms_product setelah pemrosesan: {len(cms_product_df)}")
            
            # Buat dictionary untuk memetakan SKU Bundle ke nama produk
            sku_to_name = {}
            if name_bundle_column is not None:
                for idx, row in df.iterrows():
                    sku = row["SKU Bundle"]
                    name = row[name_bundle_column] if pd.notna(row[name_bundle_column]) else "Bundle Product"
                    sku_to_name[sku] = name
            
            # Pastikan product_name dan product_desc sesuai dengan SKU Bundle
            if sku_to_name:
                cms_product_df["product_name"] = cms_product_df["product_code"].map(lambda x: sku_to_name.get(x, "Bundle Product"))
                cms_product_df["product_desc"] = cms_product_df["product_code"].map(lambda x: sku_to_name.get(x, "Bundle Product"))
            
            # Setelah pemetaan kolom
            df.columns = df.columns.str.replace(' ', '').str.strip()
            
            # Setelah membuat result_df dan cms_product_df
            date_columns_result = ["StartDate", "EndDate", "CreatedDate", "UpdatedDate"]
            date_columns_cms = ["created_at", "updated_at"]
            
            result_df = self.convert_date_columns_to_datetime(result_df, date_columns_result)
            cms_product_df = self.convert_date_columns_to_datetime(cms_product_df, date_columns_cms)
            
            return {"cms_bundle": result_df, "cms_product": cms_product_df}, df["Client"].iloc[0]
            
        except Exception as e:
            error_msg = f"Kesalahan saat memproses bundle: {str(e)}"
            self.log(error_msg, "error")
            logging.error(f"Error dalam process_bundle: {str(e)}")
            raise
    
    def save_output(self, all_sheets_data, client_name):
        """Save processed data to output file"""
        try:
            # Get values from UI
            created_by = self.created_by
            
            # Format tanggal untuk penamaan file (ddmmyyyy)
            current_date = datetime.today().strftime('%d%m%Y')
            process_type = self.process_type
            
            # Format nama client (uppercase dan hilangkan spasi berlebih)
            client_name = client_name.strip().upper()
            
            # Create file name based on process type (uppercase)
            if process_type == "Bundle":
                output_prefix = f"SETUP_BUNDLE_{client_name}_{current_date}"
            elif process_type == "Supplementary":
                output_prefix = f"SETUP_SUPPLEMENTARY_{client_name}_{current_date}"
            elif process_type == "Gift":
                output_prefix = f"SETUP_GIFT_{client_name}_{current_date}"
            
            output_format = self.output_format
            output_file_name = f"{output_prefix}.{output_format}"
            output_dir = os.path.join(os.path.dirname(self.file_path), "output")
            
            # Create output directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            output_file = os.path.join(output_dir, output_file_name)
            
            if output_format == "xlsx":
                # Create new workbook
                wb = Workbook()
                
                # Save each DataFrame to different sheets
                for sheet_name, output_df in all_sheets_data.items():
                    if isinstance(output_df, dict):
                        # Handle nested dictionaries (like Gift with Header and Line)
                        for sub_sheet_name, df in output_df.items():
                            # Ganti nilai pd.NA dengan None untuk kompatibilitas Excel
                            df = df.replace({pd.NA: None})
                            
                            # Membuat salinan df_copy untuk mendeteksi kolom highlight
                            df_copy = df.copy()
                            
                            # Cek apakah ada kolom highlight yang perlu diproses
                            highlight_columns = [col for col in df.columns if col.endswith('_highlight')]
                            
                            # Hapus kolom highlight dari df_copy yang akan ditulis ke Excel
                            for col in highlight_columns:
                                original_col = col.replace('_highlight', '')
                                # Pastikan kolom asli dan kolom highlight keduanya ada
                                if original_col in df.columns:
                                    df = df.drop(columns=[col])
                            
                            ws = wb.create_sheet(title=f"{sub_sheet_name} ({len(df)})")
                            
                            # Write headers
                            headers = df.columns.tolist()
                            for col_idx, header in enumerate(headers, start=1):
                                ws.cell(row=1, column=col_idx, value=header)
                                
                            # Write data with highlighting
                            for row_idx, row in enumerate(df.values, start=2):
                                for col_idx, value in enumerate(row, start=1):
                                    col_name = headers[col_idx - 1]
                                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                                    
                                    # Tentukan apakah cell perlu di-highlight berdasarkan kolom highlight
                                    highlight_col = f"{col_name}_highlight"
                                    if highlight_col in df_copy.columns:
                                        highlight_color = df_copy.iloc[row_idx - 2][highlight_col]
                                        if highlight_color == "yellow":
                                            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                                        elif highlight_color == "red":
                                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                                    
                                    # Highlight standar seperti sebelumnya
                                    if value == "Unknown":
                                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
                                    elif value == "Format Tanggal Salah":
                                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red highlight
                    else:
                        # Handle regular DataFrames
                        # Ganti nilai pd.NA dengan None untuk kompatibilitas Excel
                        output_df = output_df.replace({pd.NA: None})
                        
                        ws = wb.create_sheet(title=f"{sheet_name} ({len(output_df)})")
                        
                        # Write headers
                        headers = output_df.columns.tolist()
                        for col_idx, header in enumerate(headers, start=1):
                            ws.cell(row=1, column=col_idx, value=header)
                            
                        # Write data with highlighting
                        for row_idx, row in enumerate(output_df.values, start=2):
                            for col_idx, value in enumerate(row, start=1):
                                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                                if value == "Unknown":
                                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
                                elif value == "Format Tanggal Salah":
                                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red highlight
                
                # Remove default sheet if exists
                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                
                # Save workbook
                wb.save(output_file)
                
            elif output_format == "csv":
                # Save each DataFrame to separate CSV files
                for sheet_name, output_df in all_sheets_data.items():
                    if isinstance(output_df, dict):
                        for sub_sheet_name, df in output_df.items():
                            # Ganti nilai pd.NA dengan None untuk kompatibilitas
                            df = df.replace({pd.NA: None})
                            
                            # Hapus kolom highlight dari df yang akan ditulis ke CSV
                            highlight_columns = [col for col in df.columns if col.endswith('_highlight')]
                            for col in highlight_columns:
                                if col in df.columns:
                                    df = df.drop(columns=[col])
                            
                            output_csv_file = os.path.join(
                                output_dir, 
                                f"{output_prefix} - {sub_sheet_name}.csv"
                            )
                            df.to_csv(output_csv_file, index=False)
                    else:
                        # Ganti nilai pd.NA dengan None untuk kompatibilitas
                        output_df = output_df.replace({pd.NA: None})
                        
                        output_csv_file = os.path.join(
                            output_dir, 
                            f"{output_prefix} - {sheet_name}.csv"
                        )
                        output_df.to_csv(output_csv_file, index=False)
            
            self.log(f"File berhasil disimpan: {output_file_name}")
            return output_file
            
        except Exception as e:
            self.log(f"Error dalam save_output: {str(e)}", "error")
            logging.error(f"Error in save_output: {str(e)}")
            raise 

    def convert_date_format(self, date_value, is_end_date=False):
        """Convert date string to standard format"""
        # Gunakan fungsi dari utils.py
        return convert_date_format(date_value, is_end_date) 

    def process_supplementary(self, df, created_by="System"):
        try:
            # Debug: Cetak nama kolom sebelum pemrosesan
            self.log("Kolom asli: " + str(df.columns.tolist()))
            
            # Gabungkan header dan buat nama kolom yang sesuai
            df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
            
            # Debug: Cetak nama kolom setelah penggabungan
            self.log("Kolom setelah penggabungan: " + str(df.columns.tolist()))
            
            # Mapping nama kolom yang sesuai untuk input
            column_mapping = {
                'Client Unnamed: 0_level_1': 'Client',
                'Client Unnamed: 1_level_1': 'Client',
                'Client Unnamed: 2_level_1': 'Client',
                'Client Unnamed: 3_level_1': 'Client',
                'Client Unnamed: 4_level_1': 'Client',
                'Client Unnamed: 5_level_1': 'Client',
                'Client Unnamed: 6_level_1': 'Client',
                'Client Unnamed: 7_level_1': 'Client',
                'Client Unnamed: 8_level_1': 'Client',
                'Client Unnamed: 9_level_1': 'Client',
                'Client Unnamed: 10_level_1': 'Client',
                'Client Unnamed: 11_level_1': 'Client',
                'Client Unnamed: 12_level_1': 'Client',
                'Client Unnamed: 13_level_1': 'Client',
                'Client Unnamed: 14_level_1': 'Client',
                'Client Unnamed: 15_level_1': 'Client',
                'Client Unnamed: 16_level_1': 'Client',
                'Client Unnamed: 17_level_1': 'Client',
                'Client Unnamed: 18_level_1': 'Client',
                'Client Unnamed: 19_level_1': 'Client',
                'Client Unnamed: 20_level_1': 'Client',
                'Main SKU Unnamed: 1_level_1': 'Main SKU',
                'Main SKU Unnamed: 2_level_1': 'Main SKU',
                'Main SKU Unnamed: 3_level_1': 'Main SKU',
                'Main SKU Unnamed: 4_level_1': 'Main SKU',
                'Main SKU Unnamed: 5_level_1': 'Main SKU',
                'Main SKU Unnamed: 6_level_1': 'Main SKU',
                'Main SKU Unnamed: 7_level_1': 'Main SKU',
                'Main SKU Unnamed: 8_level_1': 'Main SKU',
                'Main SKU Unnamed: 9_level_1': 'Main SKU',
                'Main SKU Unnamed: 10_level_1': 'Main SKU',
                'Main SKU Unnamed: 11_level_1': 'Main SKU',
                'Main SKU Unnamed: 12_level_1': 'Main SKU',
                'Main SKU Unnamed: 13_level_1': 'Main SKU',
                'Main SKU Unnamed: 14_level_1': 'Main SKU',
                'Main SKU Unnamed: 15_level_1': 'Main SKU',
                'Main SKU Unnamed: 16_level_1': 'Main SKU',
                'Main SKU Unnamed: 17_level_1': 'Main SKU',
                'Main SKU Unnamed: 18_level_1': 'Main SKU',
                'Main SKU Unnamed: 19_level_1': 'Main SKU',
                'Main SKU Unnamed: 20_level_1': 'Main SKU',
                'Gift SKU Unnamed: 1_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 2_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 3_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 4_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 5_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 6_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 7_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 8_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 9_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 10_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 11_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 12_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 13_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 14_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 15_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 16_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 17_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 18_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 19_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 20_level_1': 'Gift SKU',
                'Gift Qty Unnamed: 1_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 2_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 3_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 4_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 5_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 6_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 7_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 8_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 9_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 10_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 11_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 12_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 13_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 14_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 15_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 16_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 17_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 18_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 19_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 20_level_1': 'Gift Qty',
                'Periode Start_Date': 'Start_Date',
                'Periode End_Date': 'End_Date',
                'Start_Date Unnamed: 1_level_1': 'Start_Date',
                'Start_Date Unnamed: 2_level_1': 'Start_Date',
                'Start_Date Unnamed: 3_level_1': 'Start_Date',
                'Start_Date Unnamed: 4_level_1': 'Start_Date',
                'Start_Date Unnamed: 5_level_1': 'Start_Date',
                'Start_Date Unnamed: 6_level_1': 'Start_Date',
                'Start_Date Unnamed: 7_level_1': 'Start_Date',
                'Start_Date Unnamed: 8_level_1': 'Start_Date',
                'Start_Date Unnamed: 9_level_1': 'Start_Date',
                'Start_Date Unnamed: 10_level_1': 'Start_Date',
                'Start_Date Unnamed: 11_level_1': 'Start_Date',
                'Start_Date Unnamed: 12_level_1': 'Start_Date',
                'Start_Date Unnamed: 13_level_1': 'Start_Date',
                'Start_Date Unnamed: 14_level_1': 'Start_Date',
                'Start_Date Unnamed: 15_level_1': 'Start_Date',
                'Start_Date Unnamed: 16_level_1': 'Start_Date',
                'Start_Date Unnamed: 17_level_1': 'Start_Date',
                'Start_Date Unnamed: 18_level_1': 'Start_Date',
                'Start_Date Unnamed: 19_level_1': 'Start_Date',
                'Start_Date Unnamed: 20_level_1': 'Start_Date',
                'End_Date Unnamed: 1_level_1': 'End_Date',
                'End_Date Unnamed: 2_level_1': 'End_Date',
                'End_Date Unnamed: 3_level_1': 'End_Date',
                'End_Date Unnamed: 4_level_1': 'End_Date',
                'End_Date Unnamed: 5_level_1': 'End_Date',
                'End_Date Unnamed: 6_level_1': 'End_Date',
                'End_Date Unnamed: 7_level_1': 'End_Date',
                'End_Date Unnamed: 8_level_1': 'End_Date',
                'End_Date Unnamed: 9_level_1': 'End_Date',
                'End_Date Unnamed: 10_level_1': 'End_Date',
                'End_Date Unnamed: 11_level_1': 'End_Date',
                'End_Date Unnamed: 12_level_1': 'End_Date',
                'End_Date Unnamed: 13_level_1': 'End_Date',
                'End_Date Unnamed: 14_level_1': 'End_Date',
                'End_Date Unnamed: 15_level_1': 'End_Date',
                'End_Date Unnamed: 16_level_1': 'End_Date',
                'End_Date Unnamed: 17_level_1': 'End_Date',
                'End_Date Unnamed: 18_level_1': 'End_Date',
                'End_Date Unnamed: 19_level_1': 'End_Date',
                'End_Date Unnamed: 20_level_1': 'End_Date',
                'Market Place Unnamed: 1_level_1': 'Marketplace',
                'Market Place Unnamed: 2_level_1': 'Marketplace',
                'Market Place Unnamed: 3_level_1': 'Marketplace',
                'Market Place Unnamed: 4_level_1': 'Marketplace',
                'Market Place Unnamed: 5_level_1': 'Marketplace',
                'Market Place Unnamed: 6_level_1': 'Marketplace',
                'Market Place Unnamed: 7_level_1': 'Marketplace',
                'Market Place Unnamed: 8_level_1': 'Marketplace',
                'Market Place Unnamed: 9_level_1': 'Marketplace',
                'Market Place Unnamed: 10_level_1': 'Marketplace',
                'Market Place Unnamed: 11_level_1': 'Marketplace',
                'Market Place Unnamed: 12_level_1': 'Marketplace',
                'Market Place Unnamed: 13_level_1': 'Marketplace',
                'Market Place Unnamed: 14_level_1': 'Marketplace',
                'Market Place Unnamed: 15_level_1': 'Marketplace',
                'Market Place Unnamed: 16_level_1': 'Marketplace',
                'Market Place Unnamed: 17_level_1': 'Marketplace',
                'Market Place Unnamed: 18_level_1': 'Marketplace',
                'Market Place Unnamed: 19_level_1': 'Marketplace',
                'Market Place Unnamed: 20_level_1': 'Marketplace',
                'Marketplace Unnamed: 0_level_1': 'Marketplace',
                'Marketplace Unnamed: 1_level_1': 'Marketplace',
                'Marketplace Unnamed: 2_level_1': 'Marketplace',
                'Marketplace Unnamed: 3_level_1': 'Marketplace',
                'Marketplace Unnamed: 4_level_1': 'Marketplace',
                'Marketplace Unnamed: 5_level_1': 'Marketplace',
                'Marketplace Unnamed: 6_level_1': 'Marketplace',
                'Marketplace Unnamed: 7_level_1': 'Marketplace',
                'Marketplace Unnamed: 8_level_1': 'Marketplace',
                'Marketplace Unnamed: 9_level_1': 'Marketplace',
                'Marketplace Unnamed: 10_level_1': 'Marketplace',
                'Marketplace Unnamed: 11_level_1': 'Marketplace',
                'Marketplace Unnamed: 12_level_1': 'Marketplace',
                'Marketplace Unnamed: 13_level_1': 'Marketplace',
                'Marketplace Unnamed: 14_level_1': 'Marketplace',
                'Marketplace Unnamed: 15_level_1': 'Marketplace',
                'Marketplace Unnamed: 16_level_1': 'Marketplace',
                'Marketplace Unnamed: 17_level_1': 'Marketplace',
                'Marketplace Unnamed: 18_level_1': 'Marketplace',
                'Marketplace Unnamed: 19_level_1': 'Marketplace',
                'Marketplace Unnamed: 20_level_1': 'Marketplace',
                'Name SKU Unnamed: 1_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 2_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 3_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 4_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 5_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 6_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 7_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 8_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 9_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 10_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 11_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 12_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 13_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 14_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 15_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 16_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 17_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 18_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 19_level_1': 'Name SKU_1',
                'Name SKU Unnamed: 20_level_1': 'Name SKU_1',
            }

            # Rename kolom
            # Membuat dictionary baru untuk case-insensitive mapping
            lowercase_column_mapping = {}
            for key, value in column_mapping.items():
                lowercase_column_mapping[key.lower()] = value
            
            # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
            rename_dict = {}
            for col in df.columns:
                col_lower = col.lower()
                if col_lower in lowercase_column_mapping:
                    rename_dict[col] = lowercase_column_mapping[col_lower]
            
            # Rename kolom menggunakan dictionary yang baru dibuat
            df = df.rename(columns=rename_dict)
            df = df.copy()
            
            # Debug: Cetak nama kolom setelah mapping
            self.log("Kolom setelah mapping: " + str(df.columns.tolist()))
            
            # Forward fill (propagate) untuk nilai yang kosong
            df.ffill(inplace=True)
            
            # Pastikan kolom yang diperlukan ada
            required_columns = ['Marketplace', 'Client', 'Start_Date', 'End_Date']
            for col in required_columns:
                if col not in df.columns:
                    raise KeyError(f"Kolom '{col}' tidak ditemukan setelah pemetaan.")
            
            # Validasi data kosong untuk supplementary
            self.validate_supplementary_columns_data(df)
            
            # Validate date ranges before processing
            self.log("Memvalidasi range tanggal untuk Supplementary...")
            invalid_date_rows = self.validate_date_range(df, 'Start_Date', 'End_Date')
            
            if invalid_date_rows:
                # Create detailed error message
                error_details = []
                for row_info in invalid_date_rows:
                    error_details.append(
                        f"Baris {row_info['row']}: Main SKU '{row_info.get('sku_bundle', 'N/A')}' "
                        f"(Client: {row_info['client']}, Marketplace: {row_info['marketplace']}) - "
                        f"Start_Date: {row_info['start_date']}, End_Date: {row_info['end_date']}"
                    )
                
                error_msg = (
                    f"Data tanggal tidak valid ditemukan di {len(invalid_date_rows)} baris. "
                    f"End_Date tidak boleh lebih awal dari Start_Date.\n\n"
                    f"Detail error:\n" + "\n".join(error_details)
                )
                
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            self.log("Validasi tanggal berhasil - semua range tanggal valid")
            
            # Split values in Marketplace column (karena bisa ada beberapa marketplace dipisahkan koma)
            df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
            df['Marketplace'] = df['Marketplace'].str.strip()
            
            # Convert date formats
            df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
            df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
            
            # Assign ShopId berdasarkan marketplace dan client
            df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
            
            # Validasi ShopId yang menghasilkan "Unknown"
            self.validate_shop_id_mapping(df)
            
            # Buat DataFrame hasil untuk format supplementary
            result_df = pd.DataFrame({
                "ItemID": df["Main SKU"],
                "Supplementary": df["Gift SKU"],
                "SupplementaryQty": df["Gift Qty"],
                "IsActive": True,
                "StartDate": df["Start_Date"],
                "EndDate": df["End_Date"],
                "ShopID": df["ShopId"],
                "CreatedAt": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
                "CreatedBy": created_by if created_by != "System" else self.created_by
            })
            
            # Setelah pemetaan kolom
            df.columns = df.columns.str.replace(' ', '').str.strip()
            
            # Konversi kolom tanggal ke datetime
            date_columns = ["StartDate", "EndDate", "CreatedAt"]
            result_df = self.convert_date_columns_to_datetime(result_df, date_columns)
            
            # Validasi hasil output untuk kolom tanggal
            self.validate_output_date_columns(result_df)
            
            return result_df, df["Client"].iloc[0]  # Return client name separately
            
        except Exception as e:
            logging.error(f"Error dalam process_supplementary: {str(e)}")
            self.log(f"Error dalam process_supplementary: {str(e)}", "error")
            raise
    
    def process_gift(self, df, created_by="System"):
        try:
            # Cek apakah kolom 'Main SKU' ada
            if 'Main SKU' not in df.columns:
                df['Main SKU'] = None  # Atur nilai menjadi NULL jika kolom tidak ada
            
            # Debug: Cetak nama kolom sebelum pemrosesan
            self.log("Kolom asli: " + str(df.columns.tolist()))
            
            # Gabungkan header dan buat nama kolom yang sesuai
            df.columns = [' '.join(str(col) for col in col_group).strip() for col_group in df.columns.values]
            
            # Debug: Cetak nama kolom setelah penggabungan
            self.log("Kolom setelah penggabungan: " + str(df.columns.tolist()))
            
            # Mapping nama kolom yang sesuai untuk input
            column_mapping = {
                'Purchase (Min -> Max) Value_Start': 'Value_Start',
                'Purchase (Min -> Max) Value_End': 'Value_End',
                'Periode Start_Date': 'Start_Date',
                'Periode End_Date': 'End_Date',
                'Client Unnamed: 0_level_1': 'Client',
                'Client Unnamed: 1_level_1': 'Client',
                'Client Unnamed: 2_level_1': 'Client',
                'Client Unnamed: 3_level_1': 'Client',
                'Client Unnamed: 4_level_1': 'Client',
                'Client Unnamed: 5_level_1': 'Client',
                'Client Unnamed: 6_level_1': 'Client',
                'Client Unnamed: 7_level_1': 'Client',
                'Client Unnamed: 8_level_1': 'Client',
                'Client Unnamed: 9_level_1': 'Client',
                'Client Unnamed: 10_level_1': 'Client',
                'Client Unnamed: 11_level_1': 'Client',
                'Client Unnamed: 12_level_1': 'Client',
                'Client Unnamed: 13_level_1': 'Client',
                'Client Unnamed: 14_level_1': 'Client',
                'Client Unnamed: 15_level_1': 'Client',
                'Client Unnamed: 16_level_1': 'Client',
                'Client Unnamed: 17_level_1': 'Client',
                'Client Unnamed: 18_level_1': 'Client',
                'Client Unnamed: 19_level_1': 'Client',
                'Client Unnamed: 20_level_1': 'Client',
                'Gift Qty Unnamed: 0_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 1_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 2_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 3_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 4_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 5_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 6_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 7_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 8_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 9_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 10_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 11_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 12_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 13_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 14_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 15_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 16_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 17_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 18_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 19_level_1': 'Gift Qty',
                'Gift Qty Unnamed: 20_level_1': 'Gift Qty',
                'Gift SKU Unnamed: 0_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 1_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 2_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 3_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 4_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 5_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 6_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 7_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 8_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 9_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 10_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 11_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 12_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 13_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 14_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 15_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 16_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 17_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 18_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 19_level_1': 'Gift SKU',
                'Gift SKU Unnamed: 20_level_1': 'Gift SKU',
                'Qty Limit Unnamed: 0_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 1_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 2_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 3_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 4_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 5_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 6_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 7_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 8_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 9_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 10_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 11_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 12_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 13_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 14_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 15_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 16_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 17_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 18_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 19_level_1': 'Limit Qty',
                'Qty Limit Unnamed: 20_level_1': 'Limit Qty',
                'Item Limit Unnamed: 0_level_1': 'Limit Qty',
                'Item Limit Unnamed: 1_level_1': 'Limit Qty',
                'Item Limit Unnamed: 2_level_1': 'Limit Qty',
                'Item Limit Unnamed: 3_level_1': 'Limit Qty',
                'Item Limit Unnamed: 4_level_1': 'Limit Qty',
                'Item Limit Unnamed: 5_level_1': 'Limit Qty',
                'Item Limit Unnamed: 6_level_1': 'Limit Qty',
                'Item Limit Unnamed: 7_level_1': 'Limit Qty',
                'Item Limit Unnamed: 8_level_1': 'Limit Qty',
                'Item Limit Unnamed: 9_level_1': 'Limit Qty',
                'Item Limit Unnamed: 10_level_1': 'Limit Qty',
                'Item Limit Unnamed: 11_level_1': 'Limit Qty',
                'Item Limit Unnamed: 12_level_1': 'Limit Qty',
                'Item Limit Unnamed: 13_level_1': 'Limit Qty',
                'Item Limit Unnamed: 14_level_1': 'Limit Qty',
                'Item Limit Unnamed: 15_level_1': 'Limit Qty',
                'Item Limit Unnamed: 16_level_1': 'Limit Qty',
                'Item Limit Unnamed: 17_level_1': 'Limit Qty',
                'Item Limit Unnamed: 18_level_1': 'Limit Qty',
                'Item Limit Unnamed: 19_level_1': 'Limit Qty',
                'Item Limit Unnamed: 20_level_1': 'Limit Qty',
                'Limit Unnamed: 0_level_1': 'Limit Qty',
                'Limit Unnamed: 1_level_1': 'Limit Qty',
                'Limit Unnamed: 2_level_1': 'Limit Qty',
                'Limit Unnamed: 3_level_1': 'Limit Qty',
                'Limit Unnamed: 4_level_1': 'Limit Qty',
                'Limit Unnamed: 5_level_1': 'Limit Qty',
                'Limit Unnamed: 6_level_1': 'Limit Qty',
                'Limit Unnamed: 7_level_1': 'Limit Qty',
                'Limit Unnamed: 8_level_1': 'Limit Qty',
                'Limit Unnamed: 9_level_1': 'Limit Qty',
                'Limit Unnamed: 10_level_1': 'Limit Qty',
                'Limit Unnamed: 11_level_1': 'Limit Qty',
                'Limit Unnamed: 12_level_1': 'Limit Qty',
                'Limit Unnamed: 13_level_1': 'Limit Qty',
                'Limit Unnamed: 14_level_1': 'Limit Qty',
                'Limit Unnamed: 15_level_1': 'Limit Qty',
                'Limit Unnamed: 16_level_1': 'Limit Qty',
                'Limit Unnamed: 17_level_1': 'Limit Qty',
                'Limit Unnamed: 18_level_1': 'Limit Qty',
                'Limit Unnamed: 19_level_1': 'Limit Qty',
                'Limit Unnamed: 20_level_1': 'Limit Qty',
                'Main SKU Unnamed: 0_level_1': 'Main SKU',
                'Main SKU Unnamed: 1_level_1': 'Main SKU',
                'Main SKU Unnamed: 2_level_1': 'Main SKU',
                'Main SKU Unnamed: 3_level_1': 'Main SKU',
                'Main SKU Unnamed: 4_level_1': 'Main SKU',
                'Main SKU Unnamed: 5_level_1': 'Main SKU',
                'Main SKU Unnamed: 6_level_1': 'Main SKU',
                'Main SKU Unnamed: 7_level_1': 'Main SKU',
                'Main SKU Unnamed: 8_level_1': 'Main SKU',
                'Main SKU Unnamed: 9_level_1': 'Main SKU',
                'Main SKU Unnamed: 10_level_1': 'Main SKU',
                'Main SKU Unnamed: 11_level_1': 'Main SKU',
                'Main SKU Unnamed: 12_level_1': 'Main SKU',
                'Main SKU Unnamed: 13_level_1': 'Main SKU',
                'Main SKU Unnamed: 14_level_1': 'Main SKU',
                'Main SKU Unnamed: 15_level_1': 'Main SKU',
                'Main SKU Unnamed: 16_level_1': 'Main SKU',
                'Main SKU Unnamed: 17_level_1': 'Main SKU',
                'Main SKU Unnamed: 18_level_1': 'Main SKU',
                'Main SKU Unnamed: 19_level_1': 'Main SKU',
                'Main SKU Unnamed: 20_level_1': 'Main SKU',
                'Market Place Unnamed: 0_level_1': 'Marketplace',
                'Market Place Unnamed: 1_level_1': 'Marketplace',
                'Market Place Unnamed: 2_level_1': 'Marketplace',
                'Market Place Unnamed: 3_level_1': 'Marketplace',
                'Market Place Unnamed: 4_level_1': 'Marketplace',
                'Market Place Unnamed: 5_level_1': 'Marketplace',
                'Market Place Unnamed: 6_level_1': 'Marketplace',
                'Market Place Unnamed: 7_level_1': 'Marketplace',
                'Market Place Unnamed: 8_level_1': 'Marketplace',
                'Market Place Unnamed: 9_level_1': 'Marketplace',
                'Market Place Unnamed: 10_level_1': 'Marketplace',
                'Market Place Unnamed: 11_level_1': 'Marketplace',
                'Market Place Unnamed: 12_level_1': 'Marketplace',
                'Market Place Unnamed: 13_level_1': 'Marketplace',
                'Market Place Unnamed: 14_level_1': 'Marketplace',
                'Market Place Unnamed: 15_level_1': 'Marketplace',
                'Market Place Unnamed: 16_level_1': 'Marketplace',
                'Market Place Unnamed: 17_level_1': 'Marketplace',
                'Market Place Unnamed: 18_level_1': 'Marketplace',
                'Market Place Unnamed: 19_level_1': 'Marketplace',
                'Market Place Unnamed: 20_level_1': 'Marketplace',
                'Marketplace Unnamed: 0_level_1': 'Marketplace',
                'Marketplace Unnamed: 1_level_1': 'Marketplace',
                'Marketplace Unnamed: 2_level_1': 'Marketplace',
                'Marketplace Unnamed: 3_level_1': 'Marketplace', 
                'Marketplace Unnamed: 4_level_1': 'Marketplace',
                'Marketplace Unnamed: 5_level_1': 'Marketplace',
                'Marketplace Unnamed: 6_level_1': 'Marketplace',
                'Marketplace Unnamed: 7_level_1': 'Marketplace',
                'Marketplace Unnamed: 8_level_1': 'Marketplace',
                'Marketplace Unnamed: 9_level_1': 'Marketplace',
                'Marketplace Unnamed: 10_level_1': 'Marketplace',
                'Marketplace Unnamed: 11_level_1': 'Marketplace',
                'Marketplace Unnamed: 12_level_1': 'Marketplace',
                'Marketplace Unnamed: 13_level_1': 'Marketplace',
                'Marketplace Unnamed: 14_level_1': 'Marketplace',
                'Marketplace Unnamed: 15_level_1': 'Marketplace',
                'Marketplace Unnamed: 16_level_1': 'Marketplace',
                'Marketplace Unnamed: 17_level_1': 'Marketplace',
                'Marketplace Unnamed: 18_level_1': 'Marketplace',
                'Marketplace Unnamed: 19_level_1': 'Marketplace',
                'Marketplace Unnamed: 20_level_1': 'Marketplace',
                'Brand Unnamed: 1_level_1': 'Client',
                'Brand Unnamed: 2_level_1': 'Client',
                'Brand Unnamed: 3_level_1': 'Client',
                'Brand Unnamed: 4_level_1': 'Client',
                'Brand Unnamed: 5_level_1': 'Client',
                'Brand Unnamed: 6_level_1': 'Client',
                'Brand Unnamed: 7_level_1': 'Client',
                'Brand Unnamed: 8_level_1': 'Client',
                'Brand Unnamed: 9_level_1': 'Client',
                'Brand Unnamed: 10_level_1': 'Client',
                'Brand Unnamed: 11_level_1': 'Client',
                'Brand Unnamed: 12_level_1': 'Client',
                'Brand Unnamed: 13_level_1': 'Client',
                'Brand Unnamed: 14_level_1': 'Client',
                'Brand Unnamed: 15_level_1': 'Client',
                'Brand Unnamed: 16_level_1': 'Client',
                'Brand Unnamed: 17_level_1': 'Client',
                'Brand Unnamed: 18_level_1': 'Client',
                'Brand Unnamed: 19_level_1': 'Client',
                'Brand Unnamed: 20_level_1': 'Client',
                'Limit Qty Unnamed: 0_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 1_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 2_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 3_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 4_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 5_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 6_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 7_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 8_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 9_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 10_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 11_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 12_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 13_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 14_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 15_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 16_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 17_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 18_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 19_level_1': 'Limit Qty',
                'Limit Qty Unnamed: 20_level_1': 'Limit Qty',            
            }
            
            # Rename kolom
            # Membuat dictionary baru untuk case-insensitive mapping
            lowercase_column_mapping = {}
            for key, value in column_mapping.items():
                lowercase_column_mapping[key.lower()] = value
            
            # Membuat dictionary untuk memetakan nama kolom saat ini ke nama kolom standar
            rename_dict = {}
            for col in df.columns:
                col_lower = col.lower()
                if col_lower in lowercase_column_mapping:
                    rename_dict[col] = lowercase_column_mapping[col_lower]
            
            # Rename kolom menggunakan dictionary yang baru dibuat
            df = df.rename(columns=rename_dict)
            df = df.copy()
            df.ffill(inplace=True)
            
            # Debug: Cetak nama kolom setelah mapping
            self.log("Kolom setelah mapping: " + str(df.columns.tolist()))
            
            # Validate date ranges before processing
            self.log("Memvalidasi range tanggal untuk Gift...")
            invalid_date_rows = self.validate_date_range(df, 'Start_Date', 'End_Date')
            
            if invalid_date_rows:
                # Create detailed error message
                error_details = []
                for row_info in invalid_date_rows:
                    error_details.append(
                        f"Baris {row_info['row']}: Main SKU '{row_info.get('sku_bundle', 'N/A')}' "
                        f"(Client: {row_info['client']}, Marketplace: {row_info['marketplace']}) - "
                        f"Start_Date: {row_info['start_date']}, End_Date: {row_info['end_date']}"
                    )
                
                error_msg = (
                    f"Data tanggal tidak valid ditemukan di {len(invalid_date_rows)} baris. "
                    f"End_Date tidak boleh lebih awal dari Start_Date.\n\n"
                    f"Detail error:\n" + "\n".join(error_details)
                )
                
                self.log(error_msg, "error")
                raise ValueError(error_msg)
            
            self.log("Validasi tanggal berhasil - semua range tanggal valid")
            
            # Validasi data kosong untuk gift
            self.validate_gift_columns_data(df)
            
            # Pisahkan nilai dalam kolom Marketplace
            if 'Marketplace' in df.columns:
                df = df.assign(Marketplace=df['Marketplace'].str.split(',')).explode('Marketplace')
                df['Marketplace'] = df['Marketplace'].str.strip()
            else:
                raise KeyError("Kolom 'Marketplace' tidak ditemukan setelah pemetaan.")
            
            # Ganti nilai "No Limit" dengan "NULL" pada kolom yang relevan
            if 'Limit Qty' in df.columns:
                df['Limit Qty'] = df['Limit Qty'].replace("No Limit", "NULL")
            
            # Assign ShopId dengan memperhitungkan Client
            df["ShopId"] = df.apply(lambda row: assign_shop_id(row["Marketplace"], row["Client"]), axis=1)
            
            # Validasi ShopId yang menghasilkan "Unknown"
            self.validate_shop_id_mapping(df)
            
            # Konversi format tanggal
            if 'Start_Date' in df.columns:
                df["Start_Date"] = df["Start_Date"].apply(lambda x: convert_date_format(x, False))
            else:
                self.log("Kolom yang tersedia: " + str(df.columns.tolist()))
                raise KeyError("Kolom 'Start_Date' tidak ditemukan")
                
            if 'End_Date' in df.columns:
                df["End_Date"] = df["End_Date"].apply(lambda x: convert_date_format(x, True))
            
            # Simpan urutan asli dari file input
            df['original_index'] = df.index
            
            # Generate GIFT ID tanpa groupby (sequential number)
            df['GIFT_ID'] = range(1, len(df) + 1)
            
            # Tambahkan counter untuk setiap kombinasi marketplace dan client
            df['MARKET_COUNTER'] = df.groupby(['Marketplace', 'Client']).cumcount() + 1
            
            # Generate FORMATTED_GIFT_ID
            df['FORMATTED_GIFT_ID'] = df.apply(create_gift_id, axis=1)
            
            # Buat DataFrame untuk Header dengan urutan kolom yang ditentukan
            header_columns = [
                "ID", "GIFT ID", "CLIENTID", "SHOPID", "GIFTTYPE", "STARTDATE", 
                "ENDDATE", "ISACTIVE", "CREATED ON", "UPDATE ON", "CREATED BY", 
                "UPDATE DATE", "LIMIT SUMMARY", "USAGE SUMMARY"
            ]
            
            line_columns = [
                "GIFT ID", "MAINSKU", "VALUESTART", "VALUEEND", "GIFTSKU", 
                "GIFTQTY", "GIFTLINENUMBER", "ISACTIVE", "MULTIPLE", 
                "ITEMLIMIT", "ITEMUSAGE"
            ]
            
            header_df = pd.DataFrame({
                "ID": None,
                "GIFT ID": df['FORMATTED_GIFT_ID'],  # Langsung gunakan FORMATTED_GIFT_ID
                "CLIENTID": df['ShopId'].apply(get_client_id),
                "SHOPID": df['ShopId'],
                "GIFTTYPE": df['Main SKU'].fillna('').apply(determine_gift_type),
                "STARTDATE": df['Start_Date'],
                "ENDDATE": df['End_Date'],
                "ISACTIVE": True,
                "CREATED ON": datetime.now().strftime("%m/%d/%Y %H:%M:%S"),
                "UPDATE ON": "NULL",
                "CREATED BY": created_by if created_by != "System" else self.created_by,
                "UPDATE DATE": "NULL",
                "LIMIT SUMMARY": df.apply(lambda row: row['Limit Qty'] if determine_gift_type(row.get('Main SKU', '')) == "2" else None, axis=1),
                "USAGE SUMMARY": "NULL"
            })[header_columns]
            
            # Simpan indeks asli sebelum pengurutan
            header_df['original_index'] = df['original_index']
            
            # Urutkan berdasarkan indeks asli
            header_df = header_df.sort_values(by=['original_index'])
            
            # Hapus kolom indeks asli setelah pengurutan
            header_df = header_df.drop(columns=['original_index'])
            
            # Buat DataFrame untuk Line dengan urutan kolom yang ditentukan
            line_df = pd.DataFrame({
                "GIFT ID": None,
                "MAINSKU": df['Main SKU'].fillna(''),
                "VALUESTART": df['Value_Start'],
                "VALUEEND": df['Value_End'],
                "GIFTSKU": df['Gift SKU'],
                "GIFTQTY": df['Gift Qty'],
                "GIFTLINENUMBER": None,
                "ISACTIVE": True,
                "MULTIPLE": False,
                "ITEMLIMIT": df.apply(lambda row: None if determine_gift_type(row.get('Main SKU', '')) == "2" else row['Limit Qty'], axis=1),
                "ITEMUSAGE": "NULL"
            })[line_columns]
            
            # Simpan indeks asli sebelum pengurutan
            line_df['original_index'] = df['original_index']
            
            # Urutkan berdasarkan indeks asli
            line_df = line_df.sort_values(by=['original_index'])
            
            # Hapus kolom indeks asli setelah pengurutan
            line_df = line_df.drop(columns=['original_index'])
            
            # Highlight GIFT ID yang kosong dan GIFTLINENUMBER
            line_df['GIFT ID_highlight'] = line_df['GIFT ID'].apply(lambda x: "yellow" if pd.isna(x) or x == "" else None)
            line_df['GIFTLINENUMBER_highlight'] = line_df['GIFTLINENUMBER'].apply(lambda x: "yellow" if pd.isna(x) or x == "" else None)
            
            # Setelah pemetaan kolom
            df.columns = df.columns.str.replace(' ', '').str.strip()
            
            # Setelah membuat header_df dan line_df
            header_date_columns = ["STARTDATE", "ENDDATE", "CREATED ON", "UPDATE DATE"]
            header_df = self.convert_date_columns_to_datetime(header_df, header_date_columns)
            
            # Validasi hasil output untuk kolom tanggal (untuk gift, validasi header_df)
            self.validate_gift_output_date_columns(header_df)
            
            return {"Header": header_df, "Line": line_df}, df["Client"].iloc[0]  # Return client name separately
            
        except Exception as e:
            self.log(f"Error dalam process_gift: {str(e)}", "error")
            logging.error(f"Error dalam process_gift: {str(e)}")
            raise

    # Ubah semua kolom tanggal dalam DataFrame menjadi tipe datetime
    def convert_date_columns_to_datetime(self, df, date_columns):
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        return df
    
    def prepare_analytics_data(self, all_sheets_data, client_name):
        """Prepare analytics data from processed setup-request data"""
        analytics_data = []
        
        try:
            for sheet_name, sheet_data in all_sheets_data.items():
                if isinstance(sheet_data, dict) and 'Header' in sheet_data:
                    # For gift type processing
                    header_df = sheet_data['Header']
                    line_df = sheet_data['Line']
                    
                    # Combine header and line data for analytics
                    for idx, header_row in header_df.iterrows():
                        if idx < len(line_df):
                            line_row = line_df.iloc[idx]
                            
                            analytics_record = {
                                'client': client_name,  # Use client_name parameter instead of CLIENTID
                                'marketplace': self.get_marketplace_from_shop_id(header_row.get('SHOPID', '')),
                                'gift_type': header_row.get('GIFTTYPE', '3'),
                                'shop_id': header_row.get('SHOPID', ''),
                                'client_id': header_row.get('CLIENTID', '')
                            }
                            analytics_data.append(analytics_record)
                
                elif isinstance(sheet_data, dict) and ('cms_bundle' in sheet_data or 'cms_product' in sheet_data):
                    # For bundle type processing (returns dict with cms_bundle and cms_product)
                    bundle_df = sheet_data.get('cms_bundle', pd.DataFrame())
                    for idx, row in bundle_df.iterrows():
                        shop_id = row.get('ShopId', '')
                        analytics_record = {
                            'client': client_name,  # Use client_name parameter
                            'marketplace': self.get_marketplace_from_shop_id(shop_id),
                            'gift_type': '3',  # Bundle is typically gift type 3
                            'shop_id': shop_id,
                            'client_id': ''
                        }
                        analytics_data.append(analytics_record)
                
                elif isinstance(sheet_data, pd.DataFrame):
                    # For supplementary type processing
                    for idx, row in sheet_data.iterrows():
                        shop_id = row.get('ShopID', '')
                        analytics_record = {
                            'client': client_name,  # Use client_name parameter
                            'marketplace': self.get_marketplace_from_shop_id(shop_id),
                            'gift_type': '3',
                            'shop_id': shop_id,
                            'client_id': ''
                        }
                        analytics_data.append(analytics_record)
            
            return analytics_data
            
        except Exception as e:
            self.log(f"Error preparing analytics data: {str(e)}", "warning")
            return []
    
    def get_marketplace_from_shop_id(self, shop_id):
        """Extract marketplace from shop ID - query from shop_mapping table"""
        if not shop_id or pd.isna(shop_id):
            return ''
        
        try:
            import sqlite3
            from setup_request.database import get_db_path
            
            # Try to query shop_mapping table
            db_path = get_db_path('shop_mapping.db')
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            # Try by shop_id string match first
            cursor.execute(
                "SELECT marketplace FROM shop_mapping WHERE shop_id = ? LIMIT 1",
                (str(shop_id),)
            )
            result = cursor.fetchone()
            
            if result:
                conn.close()
                return result[0].lower() if result[0] else ''
            
            # If not found, try substring match as fallback
            shop_id_str = str(shop_id).lower()
            if 'shopee' in shop_id_str:
                marketplace = 'shopee'
            elif 'tiktok' in shop_id_str:
                marketplace = 'tiktok'
            elif 'tokopedia' in shop_id_str:
                marketplace = 'tokopedia'
            elif 'lazada' in shop_id_str:
                marketplace = 'lazada'
            elif 'blibli' in shop_id_str:
                marketplace = 'blibli'
            elif 'zalora' in shop_id_str:
                marketplace = 'zalora'
            elif 'jubelio' in shop_id_str:
                marketplace = 'jubelio'
            elif 'desty' in shop_id_str:
                marketplace = 'desty'
            else:
                marketplace = ''
            
            conn.close()
            return marketplace
            
        except Exception as e:
            self.log(f"Error getting marketplace from shop_id {shop_id}: {str(e)}", "warning")
            return ''